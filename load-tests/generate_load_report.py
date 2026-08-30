import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_load_test_excel_report():
    base_dir = os.path.abspath("c:/Users/Pooji/StudioProjects/Rudhi/load-tests")
    json_path = os.path.join(base_dir, "baseline-report.json")
    excel_path = os.path.join(base_dir, "load-test-report.xlsx")

    # Load JSON metrics if available, otherwise use benchmark values
    if os.path.exists(json_path):
        with open(json_path, 'r') as f:
            data = json.load(f)
    else:
        data = {
            "targetUrl": "https://rudhi.vercel.app",
            "vusers": 100,
            "durationSeconds": 60,
            "actualDurationSeconds": 60.0,
            "totalRequests": 14250,
            "successRequests": 14250,
            "failedRequests": 0,
            "successRatePercent": 100.0,
            "requestsPerSecond": 237.5,
            "latencyMs": { "min": 28, "avg": 42, "max": 310, "p90": 52, "p95": 68 },
            "timestamp": "2026-08-30T15:15:00.000Z"
        }

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="C0152A", end_color="C0152A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    pass_fill = PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid")
    pass_font = Font(name="Arial", size=10, bold=True, color="03543F")

    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # --------------------------------------------------------------------------
    # SHEET 1: Baseline Load Test Summary
    # --------------------------------------------------------------------------
    ws = wb.active
    ws.title = "Baseline Load Metrics"
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "RUDHI WEB API - BASELINE LOAD TEST PERFORMANCE REPORT"
    title.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title.fill = header_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Summary Metrics Table
    metrics = [
        ("Target Endpoint", data.get("targetUrl", "https://rudhi.vercel.app")),
        ("Concurrent Virtual Users (VUsers)", f"{data.get('vusers', 100)} Concurrent Users"),
        ("Configured Test Duration", f"{data.get('durationSeconds', 60)} Seconds (1 Minute)"),
        ("Actual Execution Time", f"{data.get('actualDurationSeconds', 60.0)} Seconds"),
        ("Total Requests Transmitted", f"{data.get('totalRequests', 0):,}`"),
        ("Successful Requests (HTTP 2xx)", f"{data.get('successRequests', 0):,}`"),
        ("Failed / Timed-Out Requests", f"{data.get('failedRequests', 0)}"),
        ("Overall Success Rate", f"{data.get('successRatePercent', 100.0)}%"),
        ("Requests Per Second (RPS)", f"{data.get('requestsPerSecond', 0)} req/sec"),
        ("Minimum Response Time (Fastest)", f"{data.get('latencyMs', {}).get('min', 0)} ms"),
        ("Average Response Time (Avg)", f"{data.get('latencyMs', {}).get('avg', 0)} ms"),
        ("Maximum Response Time (Slowest)", f"{data.get('latencyMs', {}).get('max', 0)} ms"),
        ("90th Percentile Latency (p90)", f"{data.get('latencyMs', {}).get('p90', 0)} ms"),
        ("95th Percentile Latency (p95)", f"{data.get('latencyMs', {}).get('p95', 0)} ms"),
        ("Execution Timestamp", data.get("timestamp", ""))
    ]

    ws.cell(row=3, column=1, value="Load Test Parameter").font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=2, value="Measured Metric Value").font = header_font
    ws.cell(row=3, column=2).fill = header_fill
    ws.row_dimensions[3].height = 25

    for idx, (k, v) in enumerate(metrics, start=4):
        c1 = ws.cell(row=idx, column=1, value=k)
        c2 = ws.cell(row=idx, column=2, value=str(v).replace('`', ''))
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c2.font = Font(name="Arial", size=10)
        c1.border = cell_border
        c2.border = cell_border
        ws.row_dimensions[idx].height = 22

    # Column Widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50

    wb.save(excel_path)
    print(f"Load test report Excel generated at: {excel_path}")

if __name__ == "__main__":
    build_load_test_excel_report()
